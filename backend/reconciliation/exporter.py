# backend/reconciliation/exporter.py

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO

def export_report(all_tx: pd.DataFrame) -> BytesIO:
    """
    Export reconciliation + monthly summary into an Excel BytesIO object with formatting.
    Returns BytesIO ready for Streamlit download.
    """

    # --- Ensure required columns ---
    required_cols = [
        "Date", "TransactionID", "Description", "Debit", "Credit",
        "Bank", "Account", "Classification", "PairID",
        "GL Account", "GST", "Who"
    ]
    for col in required_cols:
        if col not in all_tx.columns:
            all_tx[col] = None

    # Ensure numeric for summary
    all_tx["Debit"] = pd.to_numeric(all_tx["Debit"], errors="coerce").fillna(0.0)
    all_tx["Credit"] = pd.to_numeric(all_tx["Credit"], errors="coerce").fillna(0.0)
    all_tx["GST"] = pd.to_numeric(all_tx.get("GST", 0), errors="coerce").fillna(0.0)

    # --- Pull month and year straight from normalizer date ---
    if "Date" in all_tx.columns:
        all_tx["Month"] = all_tx["Date"].apply(lambda x: x.month if hasattr(x, "month") else None)
        all_tx["Year"] = all_tx["Date"].apply(lambda x: x.year if hasattr(x, "year") else None)

    # --- Create workbook ---
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Reconciliation"

    # Write Reconciliation sheet
    ws1.append(list(all_tx.columns))
    for row in all_tx.itertuples(index=False):
        ws1.append(list(row))

    # --- Monthly Summary Sheet ---
    ws2 = wb.create_sheet("Monthly Summary")

    if "Month" in all_tx.columns and not all_tx["Month"].isna().all():
        monthly_summary = []
        for month, group in all_tx.groupby("Month"):
            internal_count = (group["Classification"] == "Internal").sum()
            incoming_count = (group["Classification"] == "Incoming").sum()
            outgoing_count = (group["Classification"] == "Outgoing").sum()
            total_income = group.loc[group["Classification"] == "Incoming", "Credit"].sum()
            total_expense = group.loc[group["Classification"] == "Outgoing", "Debit"].sum()
            total_incoming_gst = group.loc[group["Classification"] == "Incoming", "GST"].sum()
            total_outgoing_gst = group.loc[group["Classification"] == "Outgoing", "GST"].sum()
            monthly_summary.append([
                month, internal_count, incoming_count, outgoing_count,
                total_income, total_expense, total_incoming_gst, total_outgoing_gst
            ])

        ws2.append([
            "Month", "Internal Transfers", "Incoming Count",
            "Outgoing Count", "Total Incoming Income",
            "Total Outgoing Expense", "Total Incoming GST", "Total Outgoing GST"
        ])
        for row in monthly_summary:
            ws2.append(row)

        # --- Grand Total row ---
        grand_internal = sum(r[1] for r in monthly_summary)
        grand_incoming = sum(r[2] for r in monthly_summary)
        grand_outgoing = sum(r[3] for r in monthly_summary)
        grand_income = sum(r[4] for r in monthly_summary)
        grand_expense = sum(r[5] for r in monthly_summary)
        grand_incoming_gst = sum(r[6] for r in monthly_summary)
        grand_outgoing_gst = sum(r[7] for r in monthly_summary)

        ws2.append([
            "Grand Total", grand_internal, grand_incoming, grand_outgoing,
            grand_income, grand_expense, grand_incoming_gst, grand_outgoing_gst        ])

        # Highlight Grand Total row
        total_row = ws2.max_row
        for col in range(1, ws2.max_column + 1):
            cell = ws2.cell(row=total_row, column=col)
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

    # --- Formatting Reconciliation sheet ---
    fill_internal = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_incoming = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_outgoing = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    class_col = list(all_tx.columns).index("Classification") + 1
    for row in range(2, ws1.max_row + 1):
        cls = ws1.cell(row=row, column=class_col).value
        if cls == "Internal":
            ws1.cell(row=row, column=class_col).fill = fill_internal
        elif cls in ("Outgoing"):
            ws1.cell(row=row, column=class_col).fill = fill_outgoing
        elif cls in ("Incoming"):
            ws1.cell(row=row, column=class_col).fill = fill_incoming

    # --- Save to BytesIO ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_excel_bytes(all_tx: pd.DataFrame) -> bytes:
    """Return Excel file as bytes for Streamlit download"""
    output_io = export_report(all_tx)
    return output_io.read()
